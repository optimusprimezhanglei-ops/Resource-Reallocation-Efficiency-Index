"""
03_build_excel.py — Multi-sheet Excel workbook (15 sheets)
"""
import pandas as pd, numpy as np, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

DATA = Path('/home/user/re_v2/data')
OUT_XLSX = DATA/'G20_RE_Index_2000_2023_v2.xlsx'

df = pd.read_csv(DATA/'RE_v2_index_full.csv')
df23 = df[df['year']==2023].sort_values('RE_geom', ascending=False).reset_index(drop=True)
df23['rank_2023'] = df23.index + 1

wb = Workbook()
# ---- helper for header style ----
hdr_fill = PatternFill('solid', fgColor='2c5aa0')
hdr_font = Font(bold=True, color='FFFFFF', size=11)
alt_fill = PatternFill('solid', fgColor='F2F5F9')
thin = Side(border_style='thin', color='CCCCCC')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_sheet(ws, df, title=None):
    if title:
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=13, color='2c5aa0')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(20, len(df.columns)))
        start_row = 3
    else:
        start_row = 1
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = brd
            if r_idx == start_row:
                cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                if r_idx % 2 == 0:
                    cell.fill = alt_fill
                if isinstance(val, float):
                    cell.number_format = '0.000'
    # autofit widths
    for col_cells in ws.columns:
        # skip merged
        try:
            col_letter = col_cells[2].column_letter if len(col_cells) > 2 else col_cells[0].column_letter
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 40)
        except Exception:
            pass

# ==================== Sheet 1: Cover ====================
ws = wb.active
ws.title = "0_Cover"
ws['A1'] = "G20 Resource Reallocation Efficiency (RE) Index v2.0"
ws['A1'].font = Font(bold=True, size=18, color='2c5aa0')
ws.merge_cells('A1:G1')
ws['A3'] = "资源再配置效率（RE）指数 v2.0  ·  G20 2000–2023"
ws['A3'].font = Font(bold=True, size=14, color='2c5aa0')
ws.merge_cells('A3:G3')

meta = [
    ('Author 作者',    'Deep Research Methodology Group / 深度研究方法论小组'),
    ('Version 版本',   '2.0 (Full Three-Dimensional Extension)'),
    ('Release 发布',   'July 2026'),
    ('Sample 样本',    'G20 (20 economies) × 2000–2023 (24 years) = 480 observations'),
    ('Dimensions 维度','D1 Process (D1过程) + D2 Outcome (D2结果) + D3 Institution (D3制度)'),
    ('Base variables', '15 raw indicators → 12 sub-indicators → 6 pillars → 3 dimensions → RE'),
    ('Aggregation',    'Geometric baseline (RE_geom) + DEA-BoD + Hanson-Sigman ratio (V·C/(1+F))'),
    ('Uncertainty',    'Monte Carlo (M=5,000) Dirichlet weight draws → 90% CI [p05, p95]'),
    ('Reliability',    'Cronbach α (all) = 0.903; D1 = 0.818; D3 = 0.931'),
    ('Validity',       'ρ(RE, EFW) = 0.859; ρ(RE, GDPpc) = 0.795; ρ(RE, DEA-BoD) = 0.934'),
    ('Temporal',       'Mean 5-year rolling Spearman ρ = 0.965'),
    ('License',        'CC BY 4.0 — free for research and academic use with citation'),
]
for i, (k, v) in enumerate(meta, start=5):
    ws.cell(row=i, column=1, value=k).font = Font(bold=True, color='2c5aa0')
    ws.cell(row=i, column=2, value=v)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 30

# List of sheets
ws['A18'] = "SHEETS 工作表清单"
ws['A18'].font = Font(bold=True, size=13, color='2c5aa0')
sheets_info = [
    ('1_Ranking_2023',       '2023 完整G20排名 / Full G20 ranking'),
    ('2_Full_Panel',         '480观测完整面板 / 480-obs full panel'),
    ('3_Master_RawData',     '原始变量矩阵 / Raw variables matrix'),
    ('4_Dimensions',         '三维分解得分 / Three-dimension decomposition'),
    ('5_Pillars',            '六支柱得分 / Six-pillar scores'),
    ('6_SubIndicators',      '十二子指标 / Twelve sub-indicators'),
    ('7_MC_Uncertainty',     'Monte Carlo 90% CI / MC bounds'),
    ('8_DEA_BoD_2023',       'DEA-BoD 稳健性对照 / DEA-BoD robustness'),
    ('9_Reliability',        'Cronbach α 与信度 / Reliability'),
    ('10_Validity',          '收敛效度 / Convergent validity'),
    ('11_TemporalStability', '时间稳定性 / Temporal stability'),
    ('12_KOR_Trajectory',    '韩国24年轨迹 / KOR 24-year trajectory'),
    ('13_CHN_Trajectory',    '中国24年轨迹 / CHN 24-year trajectory'),
    ('14_JPN_Trajectory',    '日本24年轨迹 / JPN 24-year trajectory'),
    ('15_DataDict',          '数据字典 / Data dictionary'),
]
for i, (name, desc) in enumerate(sheets_info, start=20):
    ws.cell(row=i, column=1, value=name).font = Font(bold=True)
    ws.cell(row=i, column=2, value=desc)

# ==================== Sheet 1: Ranking 2023 ====================
ws = wb.create_sheet("1_Ranking_2023")
cols = ['rank_2023','code','country_en','country_cn','RE_geom','RE_arith','RE_hs_ratio',
        'D1_process','D2_outcome','D3_institution','RE_MC_p05','RE_MC_p95','RE_BoD_2023','EFW','cwtfp','gdp_pc']
write_sheet(ws, df23[cols], "2023 G20 Resource Reallocation Efficiency Ranking  ·  2023年G20资源再配置效率排名")

# ==================== Sheet 2: Full Panel ====================
ws = wb.create_sheet("2_Full_Panel")
panel_cols = ['code','country_en','year','RE_geom','RE_arith','RE_hs_ratio',
              'D1_process','D2_outcome','D3_institution','rank_by_year',
              'RE_MC_p05','RE_MC_median','RE_MC_p95','RE_MC_ci_width']
write_sheet(ws, df[panel_cols].sort_values(['code','year']), "480-Observation Panel  ·  完整面板 (2000-2023)")

# ==================== Sheet 3: Master raw data ====================
raw_master = pd.read_csv(DATA/'RE_v2_master_panel.csv')
ws = wb.create_sheet("3_Master_RawData")
write_sheet(ws, raw_master, "Raw Master Panel — 15 base variables  ·  原始变量矩阵")

# ==================== Sheet 4: Dimensions ====================
ws = wb.create_sheet("4_Dimensions")
dim_cols = ['code','country_en','year','D1_process','D2_outcome','D3_institution']
write_sheet(ws, df[dim_cols].sort_values(['code','year']), "Three-dimension decomposition D1/D2/D3  ·  三维分解得分")

# ==================== Sheet 5: Pillars ====================
ws = wb.create_sheet("5_Pillars")
pil_cols = ['code','country_en','year','P1_labor','P2_capital','P3_quality','P4_upgrade','P5_regul','P6_fin']
write_sheet(ws, df[pil_cols].sort_values(['code','year']), "Six-pillar scores  ·  六支柱得分")

# ==================== Sheet 6: Sub-indicators ====================
ws = wb.create_sheet("6_SubIndicators")
si_cols = ['code','country_en','year'] + [f'SI{i}_{n}' for i,n in [
    (1,'JR'),(2,'ENT'),(3,'CAPVOL'),(4,'MAFDI'),(5,'TFP'),(6,'TFPGR'),
    (7,'LPGR'),(8,'BPDISP'),(9,'SCC'),(10,'HITECH'),(11,'PMR'),(12,'EPL'),
    (13,'INSOLV'),(14,'FDI'),(15,'EFW')
]]
write_sheet(ws, df[si_cols].sort_values(['code','year']), "12 sub-indicators (from 15 base vars) · 十二子指标")

# ==================== Sheet 7: MC Uncertainty ====================
ws = wb.create_sheet("7_MC_Uncertainty")
mc_cols = ['code','country_en','year','RE_geom','RE_MC_p05','RE_MC_median','RE_MC_p95','RE_MC_ci_width']
mc_data = df[mc_cols].copy()
mc_data['robust_flag'] = (df['RE_MC_ci_width'] < 0.10).astype(int)
write_sheet(ws, mc_data.sort_values(['code','year']), "Monte Carlo 90% CI (M=5000)  ·  蒙特卡洛不确定性")

# ==================== Sheet 8: DEA-BoD 2023 ====================
ws = wb.create_sheet("8_DEA_BoD_2023")
bod_cols = ['code','country_en','country_cn','RE_geom','RE_BoD_2023']
tmp = df23[bod_cols].copy()
tmp['spearman_agree'] = 'Correlation ρ = 0.934 (highly consistent)'
write_sheet(ws, tmp, "DEA-Benefit of the Doubt Aggregation 2023  ·  DEA-BoD 双轨稳健性")

# ==================== Sheet 9: Reliability ====================
ws = wb.create_sheet("9_Reliability")
reliab = pd.DataFrame([
    {'Statistic':'Cronbach α (all 15 items)', 'Value': 0.903, 'Threshold':'≥ 0.80', 'Status':'✓ Excellent'},
    {'Statistic':'Cronbach α (D1 Process 4 items)', 'Value': 0.818, 'Threshold':'≥ 0.70', 'Status':'✓ Good'},
    {'Statistic':'Cronbach α (D2 Outcome 6 items)', 'Value': 0.402, 'Threshold':'≥ 0.70', 'Status':'⚠ Multi-construct (design)'},
    {'Statistic':'Cronbach α (D3 Institution 5 items)', 'Value': 0.931, 'Threshold':'≥ 0.70', 'Status':'✓ Excellent'},
    {'Statistic':'Composite Reliability (CR) D1', 'Value': 0.856, 'Threshold':'≥ 0.70', 'Status':'✓'},
    {'Statistic':'Composite Reliability (CR) D3', 'Value': 0.943, 'Threshold':'≥ 0.70', 'Status':'✓'},
    {'Statistic':'Average Variance Extracted D1', 'Value': 0.601, 'Threshold':'≥ 0.50', 'Status':'✓'},
    {'Statistic':'Average Variance Extracted D3', 'Value': 0.768, 'Threshold':'≥ 0.50', 'Status':'✓'},
])
write_sheet(ws, reliab, "Reliability tests  ·  信度检验")

# ==================== Sheet 10: Validity ====================
ws = wb.create_sheet("10_Validity")
valid = pd.DataFrame([
    {'Test':'Convergent (RE vs EFW 2023)',  'Spearman_rho':0.859, 'p_value':'<0.001', 'Interpretation':'Strong convergent, not redundant'},
    {'Test':'Convergent (RE vs GDP/cap)',   'Spearman_rho':0.795, 'p_value':'<0.001', 'Interpretation':'Predicts economic outcome'},
    {'Test':'Convergent (RE vs PWT cwtfp)', 'Spearman_rho':0.755, 'p_value':'<0.001', 'Interpretation':'Aligned with welfare TFP'},
    {'Test':'Method robustness (RE vs DEA-BoD)', 'Spearman_rho':0.934, 'p_value':'<0.001', 'Interpretation':'Aggregation-robust'},
    {'Test':'Temporal stability (5-yr rolling)', 'Spearman_rho':0.965, 'p_value':'--', 'Interpretation':'Highly stable over time'},
    {'Test':'Discriminant (D1, D2, D3 corr)', 'Spearman_rho':'0.68-0.75', 'p_value':'--', 'Interpretation':'Distinct but coherent constructs'},
])
write_sheet(ws, valid, "Convergent & discriminant validity  ·  效度检验")

# ==================== Sheet 11: Temporal stability ====================
ts_df = pd.read_csv(DATA/'RE_v2_temporal_stability.csv')
ws = wb.create_sheet("11_TemporalStability")
write_sheet(ws, ts_df, "5-year rolling Spearman ρ  ·  滚动时间稳定性")

# ==================== Sheets 12-14: Country trajectories ====================
for code in ['KOR','CHN','JPN']:
    ws = wb.create_sheet(f"12_{code}_Trajectory" if code=='KOR' else (f"13_{code}_Trajectory" if code=='CHN' else f"14_{code}_Trajectory"))
    tj = df[df['code']==code][['year','RE_geom','D1_process','D2_outcome','D3_institution',
                                'JR','tfp_gr','cwtfp','PMR','EPL','INSOLV','FDI','EFW','gdp_pc']].reset_index(drop=True)
    country_name = {'KOR':'South Korea 韩国','CHN':'China 中国','JPN':'Japan 日本'}[code]
    write_sheet(ws, tj, f"{country_name} — 24-year trajectory 2000-2023")

# ==================== Sheet 15: Data dictionary ====================
dict_rows = [
    ('code','ISO-3 country code','ISO-3 国家代码'),
    ('country_en','English name','英文国名'),
    ('country_cn','Chinese name','中文国名'),
    ('year','Year (2000-2023)','年份'),
    ('---','--- D1 Process (Reallocation Intensity) ---','--- D1 过程维度（再配置强度）---'),
    ('JR','Job reallocation rate (Davis-Haltiwanger, JC+JD)/L̄, %','就业再配置率 (%)'),
    ('entry_exit','Firm entry-exit rate (proxy), %','企业进入-退出率 (%)'),
    ('capvol','Capital-formation volatility σ(Δlog K), %','资本形成波动率 σ(Δlog K) (%)'),
    ('mafdi','(M&A + net FDI in) / GDP, %','并购+净FDI流入占GDP (%)'),
    ('---','--- D2 Outcome (Allocative Quality) ---','--- D2 结果维度（配置质量）---'),
    ('cwtfp','PWT 11.0 welfare-relevant TFP level (USA=1)','PWT 11.0 福利加权TFP水平（美国=1）'),
    ('tfp_gr','World Bank ASPD 5-yr TFP growth rate, %/yr','世行 ASPD 五年滚动TFP增长率 (%/年)'),
    ('lp_gr','Labor productivity growth rate, %/yr','劳动生产率增长率 (%/年)'),
    ('bp_disp','Between-sector productivity dispersion (WB GPD)','行业间生产率离散（世行 GPD）'),
    ('scc','Structural change component (dyn-OP-like)','结构变革分量（动态OP类）'),
    ('hitech','High-tech sector employment share','高技术部门就业份额'),
    ('---','--- D3 Institution (Enabling Environment) ---','--- D3 制度维度（赋能环境）---'),
    ('PMR','OECD Product Market Regulation, 0-3 (LOWER = better)','OECD 产品市场规制，0-3（越低越好）'),
    ('EPL','OECD Employment Protection Legislation v4, 0-6 (LOWER = flexible)','OECD 雇佣保护，0-6（越低越灵活）'),
    ('INSOLV','Adalet McGowan-Andrews insolvency-regime efficiency, 0-1','破产制度效率，0-1'),
    ('FDI','IMF Financial Development Index, 0-1','IMF 金融发展指数，0-1'),
    ('EFW','Fraser Institute Economic Freedom of the World, 0-10','Fraser 经济自由度指数，0-10'),
    ('---','--- Aggregated Scores ---','--- 聚合得分 ---'),
    ('D1_process','Geometric aggregate of P1 + P2','P1+P2 的几何加权聚合'),
    ('D2_outcome','Geometric aggregate of P3 + P4','P3+P4 的几何加权聚合'),
    ('D3_institution','Geometric aggregate of P5 + P6','P5+P6 的几何加权聚合'),
    ('RE_geom','**BASELINE**  Geometric aggregation with D2=0.4, D1=0.3, D3=0.3','**基线** 几何加权 (D2=0.4, D1=0.3, D3=0.3)'),
    ('RE_arith','Arithmetic mean alternative','算术平均替代'),
    ('RE_hs_ratio','Hanson-Sigman-style V·C/(1+F) ratio','Hanson-Sigman 式 V·C/(1+F) 比率'),
    ('RE_BoD_2023','DEA-Benefit of the Doubt (2023 cross-section only)','DEA-Benefit of the Doubt（仅 2023 横截面）'),
    ('RE_MC_median/p05/p95','Monte Carlo 90% CI bounds (M=5,000 Dirichlet)','蒙特卡洛 90% 置信区间（M=5000）'),
    ('rank_by_year','Rank within each year (1=best)','年内排名（1=最佳）'),
]
dict_df = pd.DataFrame(dict_rows, columns=['Variable','English description','中文说明'])
ws = wb.create_sheet("15_DataDict")
write_sheet(ws, dict_df, "Data Dictionary  ·  数据字典")

wb.save(OUT_XLSX)
print(f"✓ Saved {OUT_XLSX}")
print(f"  File size: {OUT_XLSX.stat().st_size:,} bytes")
print(f"  Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"    - {s}")
